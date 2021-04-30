# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 11:08:27 2020

@author: Meryll
"""

from datetime import datetime, timedelta
import openpyxl
import pandas as pd

import dtr
import ewibulletin
import ewisms
import measreminder
import raininfo
import webrelease


def format_cells():
    wbook=openpyxl.load_workbook("output/monitoring_ipr.xlsx")
    for sheet_name in pd.read_excel('output/monitoring_ipr.xlsx', sheet_name=None).keys():
        sheet=wbook[sheet_name]
        sheet.merge_cells('A2:E2')
        sheet.merge_cells('C3:E3')
        sheet.merge_cells('A4:A11')
        sheet.merge_cells('B4:B11')
        sheet.merge_cells('C4:C9')
        sheet.merge_cells('C10:C11')
        sheet.merge_cells('C12:E12')
        sheet.merge_cells('A13:A20')
        sheet.merge_cells('C13:E13')
        sheet.merge_cells('C14:E14')
        sheet.merge_cells('C15:E15')
        sheet.merge_cells('C16:E16')
        sheet.merge_cells('C17:E17')
        sheet.merge_cells('C18:E18')
        sheet.merge_cells('C19:E19')
        sheet.merge_cells('C20:E20')
        sheet.merge_cells('A21:A31')
        sheet.merge_cells('B23:B28')
        sheet.merge_cells('C21:E21')
        sheet.merge_cells('C22:E22')
        sheet.merge_cells('C23:C28')
        sheet.merge_cells('C29:E29')
        sheet.merge_cells('C30:E30')
        sheet.merge_cells('C31:E31')
        sheet.merge_cells('A32:B37')
        sheet.merge_cells('C32:C35')
        sheet.merge_cells('C36:C37')    
    wbook.save("output/monitoring_ipr.xlsx")
    

def timeline(year, quarter):
    if quarter == 4:
        end = "{}-01-01".format(year+1)
    else:
        end = "{}-{}-01".format(year, str(3*quarter + 1).zfill(2))
    end = pd.to_datetime(end)
    start = pd.to_datetime("{}-{}-01".format(year, str(3*quarter - 2).zfill(2)))
    return start, end

def get_sheet(key, sheet_name):
    url = 'https://docs.google.com/spreadsheets/d/{key}/gviz/tq?tqx=out:csv&sheet={sheet_name}&headers=1'.format(
        key=key, sheet_name=sheet_name.replace(' ', '%20'))
    df = pd.read_csv(url)
    df = df.drop([col for col in df.columns if col.startswith('Unnamed')], axis=1)
    return df

def get_shift(key, sheet_name):
    df = get_sheet(key, sheet_name)
    df = df.drop([col for col in df.columns if col.startswith('Unnamed')], axis=1)
    df.loc[:, 'Date'] = pd.to_datetime(df.loc[:, 'Date'].ffill())
    df.loc[:, 'Shift'] = pd.to_timedelta(df.loc[:, 'Shift'].map({'AM': timedelta(hours=8), 'PM': timedelta(hours=20)}))
    df.loc[:, 'ts'] = pd.to_datetime(df.loc[:, 'Date'] + df.loc[:, 'Shift'])
    return df.loc[:, ['ts', 'IOMP-MT', 'IOMP-CT']]


def main(start, end, update_existing=True, update_dtr=True, recompute=True, mysql=True):
    key = "1UylXLwDv1W1ukT4YNoUGgHCHF-W8e3F8-pIg1E024ho"
    date_range = pd.date_range(start=start, end=end, freq='M', closed='left')
    shift_sched = pd.DataFrame()
    for ts in date_range:
        sheet_name = ts.strftime('%B %Y')
        shift_sched = shift_sched.append(get_shift(key, sheet_name))
    shift_sched = shift_sched.loc[(shift_sched.ts > start) & (shift_sched.ts < end)]
    
    if update_existing:
        monitoring_ipr = pd.read_excel('output/monitoring_ipr.xlsx', sheet_name=None)
        writer = pd.ExcelWriter('output/monitoring_ipr.xlsx')
        for name in monitoring_ipr.keys():
            indiv_ipr = monitoring_ipr[name]
            indiv_ipr = indiv_ipr.loc[:, ~indiv_ipr.columns.str.contains('Unnamed:', na=False)]
            for shift_type in ['MT', 'CT']:
                curr_shift = shift_sched.loc[shift_sched['IOMP-{}'.format(shift_type)] == name, :]
                shift_list = set(curr_shift.ts) - set(map(pd.to_datetime, indiv_ipr.columns[5:]))
                for ts in shift_list:
                    indiv_ipr.loc[indiv_ipr.Category == 'Shift', str(ts)] = shift_type
            monitoring_ipr[name] = indiv_ipr

    else:
        writer = pd.ExcelWriter('output/monitoring_ipr.xlsx')
        grading_system = pd.read_excel('input/baseline.xlsx', sheet_name='format')
        monitoring_ipr = {}
    
        personnel_sheet = "personnel"    
        
        name_df = get_sheet(key, personnel_sheet)
        name_list = name_df.loc[name_df.current == 1, 'Nickname'].values
            
        for name in name_list:
            indiv_ipr = grading_system.copy()
            for shift_type in ['MT', 'CT']:
                curr_shift = shift_sched.loc[shift_sched['IOMP-{}'.format(shift_type)] == name, :]
                for ts in curr_shift.ts:
                    indiv_ipr.loc[indiv_ipr.Category == 'Shift', str(ts)] = shift_type
            monitoring_ipr[name] = indiv_ipr
    
    
    for sheet_name, xlsxdf in monitoring_ipr.items():
        xlsxdf.to_excel(writer, sheet_name, index=False)
    writer.save()

    ewisms.main(start=start, end=end, recompute=recompute)
    dtr.main(update_dtr)
    measreminder.main(start, end, mysql=mysql)
    raininfo.main(start, end, mysql=mysql)
    webrelease.main(start, end, mysql=mysql)
    ewibulletin.main(start, end, mysql=mysql)
    
    format_cells()
    
###############################################################################
if __name__ == "__main__":
    run_start = datetime.now()
    
    start = pd.to_datetime('2020-07-15')
    end = pd.to_datetime('2020-12-01')
    main(start, end, update_dtr=False)
    
    runtime = datetime.now() - run_start
    print("runtime = {}".format(runtime))